import comtypes.client
import sys
import pandas as pd
from openpyxl import Workbook

from utils import extractions


from elements.story import Story

from dxf_drawer.drawing import Drawing
from dxf_drawer.detail import Detail
from dxf_drawer.column import RectangularColumn




def create_dxf_file(column_data: list[dict]):
    pass


def generate_excel_table(stories_data, grid_lines):
    
    stories_reverse = []
    
    
    for story in stories_data[::-1]:
        stories_reverse.append(story)
        
    wb = Workbook()
    ws = wb.active
    
    # Create headers
    ws['A1'] = "NIVEL"
    ws['B1'] = "DESCRIPCION"
    
    
    for item in range(0,len(stories_reverse)-1):
        ws.cell(row=9*item+2, column=1).value = f"{stories_reverse[item]['name']}@{stories_reverse[item+1]['name']}"
        ws.cell(row=9*item+2, column=2).value = "b x h"
        ws.cell(row=9*item+3, column=2).value = "f'c"
        ws.cell(row=9*item+4, column=2).value = "As"
        ws.cell(row=9*item+5, column=2).value = "Est. en Lo"
        ws.cell(row=9*item+6, column=2).value = "Est. en Resto"
        ws.cell(row=9*item+7, column=2).value = "Estribo Externo"
        ws.cell(row=9*item+8, column=2).value = "Estribos Interno"
        ws.cell(row=9*item+9, column=2).value = "Lo"
        ws.cell(row=9*item+10, column=2).value = "Detalle"
        
        
    # Columns Data
    counter_col = 0
    for x in grid_lines:
        ws.cell(row=1, column=3+counter_col).value = f"C-{x}"
        counter_col += 1
    
        
    
    wb.save('cuadro_columnas.xlsx')

def get_story_by_elevation(stories_data, elevation):
    for story in stories_data:
        if story['elevation'] == elevation:
            return story['name']
    return None
    

def get_model_data(model_path):
    data_output = []
    
    # Create API helper object
    helper = comtypes.client.CreateObject('ETABSv1.Helper')
    helper = helper.QueryInterface(comtypes.gen.ETABSv1.cHelper)
    
    EtabsObject = helper.CreateObjectProgID("CSI.ETABS.API.ETABSObject")

    # Start ETABS application
    EtabsObject.ApplicationStart()
    # Create SnapModel Object
    SapModel =EtabsObject.SapModel
    print(SapModel)
    SapModel.File.OpenFile(model_path)
    
    # Open and save the model
    print("ETABS model opened successfully!")

    # Get the model's name to verify the connection
    ModelName = EtabsObject.SapModel.GetModelFilename()
    
    print(f"Model loaded: {ModelName}")

    print("\nFetching story information...")
    
    materials_dict = extractions.get_all_materials(SapModel)
    
    rebar_info = extractions.get_all_rebars(SapModel)
    print(rebar_info)
    
    
    stories = extractions.get_story_data(SapModel)
    print(stories)
    
    if not stories:
        print("Could not retrieve story information. Exiting.")
                # Optional: Release COM objects
                # sap_model = None
                # if 'etabs_object' in locals() and etabs_object is not None:
                #     etabs_object = None
        sys.exit()
        
    data_stories = []
    counter_stories = 0
        
    for story in stories:
        counter_stories += 1
        data_stories.append(Story(id=counter_stories, name=story['name'], elevation=story['elevation']))
        # print(f"  Story: {story['name']}, Elevation: {story['elevation']:.2f}")
        
    columns_at_levels = extractions.extract_columns_by_level(SapModel,stories)
    print(columns_at_levels)
    
    if columns_at_levels:
        print("\nExtracting columns for each level...")
        found_any_columns = False
        for story_info in stories:
           
            story_name = story_info["name"]
            if story_name in columns_at_levels and columns_at_levels[story_name]:
                found_any_columns = True
                print(f"\nLevel: {story_name} (Elevation: {story_info['elevation']:.2f})")
                for col_name in sorted(columns_at_levels[story_name]):
                    info = {}
                    col_section = SapModel.FrameObj.GetSection(col_name)[0]
                    material_defined = SapModel.PropFrame.GetMaterial(col_section)[0]
                    col_point1, col_point2,  ret_points = SapModel.FrameObj.GetPoints(col_name)
                    col_x_pos = SapModel.PointObj.GetCoordCartesian(col_point1)[0]
                    col_y_pos =SapModel.PointObj.GetCoordCartesian(col_point1)[1]
                    col_z_start = SapModel.PointObj.GetCoordCartesian(col_point1)[2]
                    col_z_end = SapModel.PointObj.GetCoordCartesian(col_point2)[2]
                    col_type_enum = SapModel.PropFrame.GetTypeOAPI(col_section)[0]
                    col_shape = None
                    if col_type_enum == 8:
                        col_shape = "Rectangular"
                        width, depth = SapModel.PropFrame.GetRectangle(col_section)[2:4]
                        
                        rebar_data = extractions.get_rebar_data(SapModel, col_section, col_name)
                    elif col_type_enum == 9:
                        col_shape = "Circular"
                        width = SapModel.PropFrame.GetCircle(col_section)[2]
                        rebar_data = extractions.get_rebar_data(SapModel, col_section,col_name)
                        # print(SapModel.PropFrame.GetRebarType(col))
                        depth =None
                    elif col_type_enum == 28:
                        col_shape = "L"
                        width = None
                        depth = None
                        rebar_data = None
                    
                    #print(f"  - {col_name} - {col_section},Material: {material_defined},  Pos X: {col_x_pos}, Pos Y: {col_y_pos}, Elev: {col_z_start} to Elev: {col_z_end}, shape: {col_shape}, Width: {width}, Depth: {depth}")
                    info['col_id'] = col_name
                    info['section'] = col_section
                    info['type'] = col_shape
                    info['width'] = width
                    info['depth'] = depth
                    info['material'] = material_defined
                    info['pos_x'] = round(col_x_pos,2)
                    info['pos_y'] = round(col_y_pos,2)
                    info['story'] = story_name
                    info['story_elevation'] = round(story_info['elevation'],2)
                    info['story_start'] = get_story_by_elevation(stories, col_z_start)
                    info['story_end'] = get_story_by_elevation(stories, col_z_end)
                    info['z_start'] = round(col_z_start,2)
                    info['z_end'] = round(col_z_end,2)
                    if rebar_data:
                        info['Mat. Rebar'] = rebar_data['mat_rebar_long']
                        info['cover'] = rebar_data['cover']
                        info['number_r2_bars'] = rebar_data['number_r2_bars']
                        info['number_r3_bars'] = rebar_data['number_r3_bars']
                        info['# Bars'] = rebar_data['number_bars']
                        info['Rebar'] = rebar_data['rebar_type']
                        info['Mat. Estribo'] = rebar_data['mat_rebar_confine']
                        
                    else:
                        info['Mat. Rebar'] = ""
                        info['cover'] = ""
                        info['number_r2_bars'] = ""
                        info['number_r3_bars'] = ""
                        info['# Bars'] = ""
                        info['Rebar'] = ""
                        info['Mat. Estribo'] = ""
                    data_output.append(info)
                    
                    
                    
                # Optionally print levels with no columns found
                # else:
                #    print(f"\nLevel: {story_name} (Elevation: {story_info['elevation']:.2f})")
                #    print("  - No columns found with top at this level.")
            
                if not found_any_columns:
                    print("\nNo columns were found associated with any story levels based on the criteria.")
            else:
                print("No column data was extracted or an error occurred.")
                
        df_columns = pd.DataFrame(data_output)
        
        # 1. Create tuple of cols "pos_x" and "pos_y" for each row.
        # This is useful to have an object that can be used to find unique combinations.
        df_columns['temp_grid'] = df_columns[['pos_x', 'pos_y']].apply(tuple, axis=1)
        
        # 2. Get unique values of these tuples and map to an integer value
        # Use factorize(), which is efficient for this. Returns
        # - an array of integers that represents the categories.
        # - an arrray of unique categories
        # Add 1 in order that identifiers start in 1 not in 0
        df_columns['GridLine'] = pd.factorize(df_columns['temp_grid'])[0] + 1
        
        grid_lines = df_columns['GridLine'].unique()
        sections = df_columns['section'].unique()
        number_details = len(sections)
        
        # 3. (Optional) Delete column temp
        df_columns = df_columns.drop(columns=['temp_grid'])
        
        #4. Sort rows
        df_sorted =df_columns.sort_values(by=['GridLine', 'story_elevation'],ascending=True)
        # Sort dataframe by pos_x, pos_y
        df_sorted.to_excel("column_output.xlsx")
                    

    #Close Application
    EtabsObject.ApplicationExit(False)
    
    # Generate Cuadro de Columnas Excel
    generate_excel_table(stories, grid_lines)
    
    # Generate dxf section details
    columns = []
    for section in sections:
        width = df_sorted[df_sorted['section'] == section].iloc[0]['width'] * 1000 # Convert to mm
        depth = df_sorted[df_sorted['section'] == section].iloc[0]['depth'] * 1000 # Convert to mm
        fc = df_sorted[df_sorted['section'] == section].iloc[0]['material']
        fc = int(fc[:-3])
        # Convert fc to kg/cm2
        fc = int(fc * (12*12)*(3.28*3.28)/(100*100*2.204))
        fc = str(fc)
        
        r2_bars = df_sorted[df_sorted['section'] == section].iloc[0]['number_r2_bars']
        r3_bars = df_sorted[df_sorted['section'] == section].iloc[0]['number_r3_bars']
        rebar_type = df_sorted[df_sorted['section'] == section].iloc[0]['Rebar']
        number_bars = df_sorted[df_sorted['section'] == section].iloc[0]['# Bars']
        cover = df_sorted[df_sorted['section'] == section].iloc[0]['cover'] * 1000 #Convert to mm
        stirrup_type = "#4"
        
        columns.append(
            RectangularColumn(width=depth, height=width, fc=fc, number_of_bars=number_bars, rebar_type=rebar_type, r2_bars=r2_bars, r3_bars=r3_bars, cover = cover, stirrup_type=stirrup_type),
        )
        
     # 1. Create list of Detail
    list_details = []
    start_point = (100,100)
    counter = 0
    width_detail = 3000
    height_detail = 3000
        
    for i in range(1, len(columns)+1): 
        actual_col = columns[i-1]
        origin_point = (start_point[0], start_point[1] - (height_detail*counter))
        detail = Detail(f"DC-{i}",origin_point, width_detail, height_detail)
        detail.set_column(actual_col)
        detail.set_origin_for_col(actual_col.width, actual_col.height)
        list_details.append( detail)
        counter += 1

    drawing = Drawing(filename='detalles_cols_etabs.dxf', list_details=list_details)
    drawing.create_dxf()
        
        
        
        
        
   